import openpyxl
from decimal import Decimal
from utils import DataClass
ids_save = []


def save(date:dict):
    cal = """Ссылка на товар
    Артикул
    Название
    Цена
    Описание
    Ссылки на изображения через запятую
    Описание
    Все характеристики с сохранением их структуры
    Название селлера
    Ссылка на селлера
    Размеры товара через запятую 
    Остатки по товару (число)
    Рейтинг
    Количество отзывов""".split('\n')
    print(len(cal))

    file_path = 'data_save.xlsx'
    file_path_filter = 'data_save_filter.xlsx'
    file_path_ids = 'data_save_ids.xlsx'
    try:
        workbook = openpyxl.load_workbook(file_path)
        workbookfilter = openpyxl.load_workbook(file_path_filter)
        workbookids = openpyxl.load_workbook(file_path_ids)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        workbookfilter = openpyxl.Workbook()
        workbookids = openpyxl.Workbook()

        sheet = workbook.active
        sheetfilter = workbookfilter.active
        sheetids = workbookids.active

        sheet.append(cal)
        sheetfilter.append(cal)
        sheetids.append(cal)

        workbook.save(file_path)
        workbookfilter.save(file_path_filter)
        workbookids.save(file_path_ids)

        workbook = openpyxl.load_workbook(file_path)
        workbookfilter = openpyxl.load_workbook(file_path_filter)
        workbookids = openpyxl.load_workbook(file_path_ids)

        sheet = workbook.active
        sheetfilter = workbookfilter.active
        sheetids = workbookids.active








    all = []
    #Ещё один костыль. Зато работает) Здесь мы получаем все наши списки карточек которые спарсили )
    for i in date:
        for j in i['card']:
            for z in j:
                    all.append(z)
    print(f'Пришло на сохранение {len(all)}')
    #print(all)
    for info in all:

        if not info.article in ids_save:
            ids_save.append(info.article)
            sheetids.append([info.article])
        else:
            print(f'Дубль XLSX, {info.article}')
            continue
        print(f"ids_save {len(ids_save)}")
        new_data = [
            str(info.urlCard),
            str(info.article),
            str(info.name),
            int(str(info.price)),
            str(info.description),
            str(info.imgs),
            str(info.description),
            str(info.opts),
            str(info.supplierName),
            str(info.url_supplier),
            str(info.sizesName),
            str(info.qty),
            Decimal(info.rating),
            str(info.feedBack),
        ]
        sheet.append(new_data)
        #доп файл. Важно на прямую float нельзя сравнивать, тем более в javascript) Используем Decimal

        # Цена
        #Страна РФ
        country = False
        if (Decimal(4.5) <= Decimal(str(info.rating))) and (str(info.opts).upper().find("россия".upper())) and (Decimal(str(info.price)) < Decimal(10000)):
            #доп фильтр и файл)
            sheetfilter.append(new_data)
    workbook.save(file_path)
    workbookfilter.save(file_path_filter)
    workbookids.save(file_path_ids)
    print(f'было сохранено: {sheet.max_row}')

